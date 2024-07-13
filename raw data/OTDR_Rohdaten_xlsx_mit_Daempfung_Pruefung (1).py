import pandas as pd
import glob
from statistics import mean
import csv
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import argparse
import re

GW_1310 = 4.95
GW_1550 = 3.93
GW_1625 = 4.2

parser = argparse.ArgumentParser(
    prog='OTDR Raw Data Checker', 
    description='The program reads raw OTDR data in XLSX format and outputs the cable length and attenuation for each address and wavelength. It also checks if the attenuations are below the threshold value. Command line arguments: python otdr_check_rd.py [--files optional_path_to_directory] [--splices number_of_splices (integer or cell reference)] [--extra additional_attenuation]',
    epilog=''
)
parser.add_argument('-f', '--files', default=os.getcwd())
parser.add_argument('-s', '--splices', default=3)
parser.add_argument('-e', '--extra', type=float, default=0.75)
args = parser.parse_args()
argv = vars(args)

try:
    path = argv["files"] 
    os.chdir(path)
    filenames = glob.glob(path + "\\*.xlsx")
except Exception as e:
    print(e, "Couldn't find file directory")
try: 
    with open("OTDR.csv", mode="w") as OTDR_file:
        OTDR_writer = csv.writer(OTDR_file, delimiter=",", lineterminator="\r")
        OTDR_writer.writerow(
            [
                "Address",
                "Cable Number",
                "Average Cable Length",
                "Deviation",
                "Cable Lengths",
                "Span Loss 1310 [dB]",
                "Span Loss 1550",
                "Span Loss 1625",
                "Home-SAI [m]",
            ]
        )
        for file in filenames:
            print(file)
            wb = load_workbook(file, data_only=True)
            sh = wb.worksheets[0]
            pipe = "None"
            try:
                if sh.cell(8, 1).value == "Cable ID":
                    pipe = sh.cell(9, 1).value
                elif sh.cell(8, 11).value == "Cable ID":
                    pipe = sh.cell(9, 11).value
            except Exception as e:
                print(e, "Couldn't read , cable ID")
            address = "None"
            try:
                address = sh.cell(9, 8).value + sh.cell(9, 12).value
            except Exception as e:
                print(e, "Couldn't read adress data")
            length = []
            span_1310 = []
            span_1550 = []
            span_1625 = []
            for x in range(1, len(wb.sheetnames)):
                sheet = wb.worksheets[x]
                cable = sheet.cell(25, 4).value
                span = sheet.cell(25, 10).value
                nm = sheet.cell(19, 1).value
                try:
                    cable_float = float(cable)
                    length.append(cable_float)
                    if "1310" in str(nm):
                        try:
                            span_1310.append(float(span))
                        except Exception as e:
                            print(e, "Couldn't read spanloss data for 1310 nm")
                    elif "1550" in str(nm):
                        try:
                            span_1550.append(float(span))
                        except Exception as e:
                            print(e, "Couldn't read spanloss data for 1550 nm")
                    elif "1625" in str(nm):
                        try:
                            span_1625.append(float(span))
                        except Exception as e:
                            print(e, "Couldn't read spanloss data for 1625 nm")
                except Exception as e:
                    print(e, "Couldn't read cable length")
            if len(span_1625) == 0:
                span_1625.append(0)
            if len(span_1550) == 0:
                span_1550.append(0)
            if len(span_1310) == 0:
                span_1310.append(0)
            OTDR_writer.writerow(
                [
                    address,
                    pipe,
                    round(mean(length), 3),
                    round(max(length) - min(length), 2),
                    length,
                    round(mean(span_1310), 3),
                    round(mean(span_1550), 3),
                    round(mean(span_1625), 3),
                ]
            )
    filenames = glob.glob(path + "\\*.csv")
    for file in filenames:
        read_file = pd.read_csv(file, encoding="latin-1")
        read_file.to_excel("OTDR_Excel.xlsx", index=None, header=True)
except Exception as e:
        print(e, "Couldn't open OTDR files or create the summary file")

try:
    wb = load_workbook("OTDR_Excel.xlsx")
    ws = wb.worksheets[0]
    extra = float(argv["extra"])
    if bool(re.compile(r"^[a-zA-Z][0-9]$").match(argv["splices"]))
        try:
            splices = int(ws.cell(argv['splices']).value)
        except Exception as e:
            print(e, "Value in given cell can't be used as amount of splices")  
    else:
        try:
            splices = int(argv["splices"])
        except Exception as e:
            print(e, "Given value for splices is not a cell or integer")
    for row in range(2, ws.max_row + 1):
        length = ws.cell(row=row, column=3).value
        GW_splice = 0.2
        GW_span_1310 = (0.36 * length + 0.45 + 0.7 + extra) + (splices * GW_splice)
        GW_span_1550 = (0.21 * length + 0.45 + 0.7 + extra) + (splices * GW_splice)
        GW_span_1625 = (0.25 * length + 0.45 + 0.7 + extra) + (splices * GW_splice)
        span_1310 = ws.cell(row=row, column=6).value
        span_1550 = ws.cell(row=row, column=7).value
        span_1625 = ws.cell(row=row, column=8).value
        if span_1310 > GW_span_1310:
            ws.cell(row=row, column=6).fill = PatternFill(
                start_color="00FF0000", fill_type="solid"
            )
        if span_1550 > GW_span_1550:
            ws.cell(row=row, column=7).fill = PatternFill(
                start_color="00FF0000", fill_type="solid"
            )
        if span_1625 > GW_span_1625:
            ws.cell(row=row, column=8).fill = PatternFill(
                start_color="00FF0000", fill_type="solid"
            )

    wb.save("OTDR_Excel_checked.xlsx")
except Exception as e:
    print(e, "Couldn't check values")

print("Done")
