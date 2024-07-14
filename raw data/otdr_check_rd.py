import pandas as pd
import glob
from statistics import mean
import csv
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import argparse
import re

# Constants for threshold values
GW_1310 = 4.95
GW_1550 = 3.93
GW_1625 = 4.2

# Argument parser setup
parser = argparse.ArgumentParser(
    prog='OTDR Raw Data Checker',
    description='Reads raw OTDR data in XLSX format and outputs cable length and attenuation for each address and wavelength. Checks if the attenuations are below the threshold value.',
)
parser.add_argument('-f', '--files', default=os.getcwd(), help='Path to the directory containing XLSX files')
parser.add_argument('-s', '--splices', type=int, default=3, help='Number of splices (integer)')
parser.add_argument('-e', '--extra', type=float, default=0.75, help='Additional attenuation')
args = parser.parse_args()
argv = vars(args)

try:
    # Set directory path
    path = argv["files"]
    os.chdir(path)
    filenames = glob.glob(path + "/*.xlsx")
except Exception as e:
    print(e, "Couldn't find file directory")
    exit(1)

# CSV file to store results
csv_filename = "OTDR.csv"
with open(csv_filename, mode="w", newline='') as OTDR_file:
    OTDR_writer = csv.writer(OTDR_file, delimiter=",")
    OTDR_writer.writerow([
        "Address", "Cable Number", "Average Cable Length", "Deviation", "Cable Lengths",
        "Span Loss 1310 [dB]", "Span Loss 1550", "Span Loss 1625", "Home-SAI [m]",
    ])

    for file in filenames:
        print(f"Processing file: {file}")
        wb = load_workbook(file, data_only=True)
        sh = wb.worksheets[2]
        pipe = "None"

        # Reading cable ID
        try:
            if "ID" in sh.cell(8, 1).value :
                pipe = sh.cell(9, 1).value
            elif "ID" in sh.cell(8, 11).value:
                pipe = sh.cell(9, 11).value
        except Exception as e:
            print(e, "Couldn't read cable ID")

        # Reading address
        try:
            addr_1 = sh.cell(13, 3).value or ""
            addr_2 = sh.cell(13, 11).value or ""
            address = f"{addr_1}, {addr_2}"
        except Exception as e:
            print(e, "Couldn't read address data")
            address = "None, None"

        length = []
        span_1310 = []
        span_1550 = []
        span_1625 = []

        # Reading cable length and span loss
        for sheet in wb.worksheets[2:]:
            try:
                cable = float(sheet.cell(25, 4).value)
                length.append(cable)
                nm = sheet.cell(19, 1).value
                span = float(sheet.cell(25, 10).value)

                if "1310" in str(nm):
                    span_1310.append(span)
                elif "1550" in str(nm):
                    span_1550.append(span)
                elif "1625" in str(nm):
                    span_1625.append(span)
            except Exception as e:
                print(e, "Couldn't read data")

        # Filling missing span losses with 0
        span_1625 = span_1625 or [0]
        span_1550 = span_1550 or [0]
        span_1310 = span_1310 or [0]

        # Writing data to CSV
        try:
            OTDR_writer.writerow([
                address, pipe, round(mean(length), 3), round(max(length) - min(length), 2), length,
                round(mean(span_1310), 3), round(mean(span_1550), 3), round(mean(span_1625), 3),
            ])
        except Exception as e:
            print(e, "Couldn't write data to CSV")

# Convert CSV to Excel
csv_df = pd.read_csv(csv_filename, encoding="latin-1")
excel_filename = "OTDR_Excel.xlsx"
csv_df.to_excel(excel_filename, index=False, header=True)

# Checking values and applying formatting
try:
    wb = load_workbook(excel_filename)
    ws = wb.active
    extra = float(argv["extra"])
    splices = int(argv["splices"])

    # Iterate over rows and apply checks
    for row in range(2, ws.max_row + 1):
        length = ws.cell(row=row, column=3).value
        GW_splice = 0.2
        GW_span_1310 = (0.36 * length + 0.45 + 0.7 + extra) + (splices * GW_splice)
        GW_span_1550 = (0.21 * length + 0.45 + 0.7 + extra) + (splices * GW_splice)
        GW_span_1625 = (0.25 * length + 0.45 + 0.7 + extra) + (splices * GW_splice)

        span_1310 = ws.cell(row=row, column=6).value
        span_1550 = ws.cell(row=row, column=7).value
        span_1625 = ws.cell(row=row, column=8).value

        if span_1310 > GW_span_1310:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="00FF0000", fill_type="solid")
        if span_1550 > GW_span_1550:
            ws.cell(row=row, column=7).fill = PatternFill(start_color="00FF0000", fill_type="solid")
        if span_1625 > GW_span_1625:
            ws.cell(row=row, column=8).fill = PatternFill(start_color="00FF0000", fill_type="solid")

    checked_excel_filename = "OTDR_Excel_checked.xlsx"
    wb.save(checked_excel_filename)
except Exception as e:
    print(e, "Couldn't check values")

# Clean up temporary files
try:
    os.remove(csv_filename)
except Exception as e:
    print(e, "Couldn't delete temporary CSV file")

try:
    os.remove(excel_filename)
except Exception as e:
    print(e, "Couldn't delete temporary Excel file")

print("Done")


