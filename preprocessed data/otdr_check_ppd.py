import pandas as pd
import glob
import csv
import os
import sys
from openpyxl import load_workbook

def main(path=path):
    # Define the path to the directory containing XLSX files (default is current working directory)
    path = os.getcwd()
    # optionally the directory to the files can be given in the command line
    if len(sys.argv) == 2:
        path = sys.argv[1]
    else if len(sys.argv) > 2:
        raise TypeError("Too many arguments. Command only has to contain the script.name and optionally the directory of the files")
    # Change the current working directory to the specified path
    os.chdir(path)
    # Print the result returned by the function print_result() with the results of cable_length() and attenuation() functions
    print(print_result(cable_length(path), attenuation(path)))

def cable_length(path):
    num_adress = 0
    # Create or open a CSV file named 'OTDR.csv' for writing
    with open('OTDR.csv', mode='w') as OTDR_file:
        # Create a CSV writer object
        OTDR_writer = csv.writer(OTDR_file, delimiter=',', lineterminator='\r')
        # Write the header row to the CSV file
        OTDR_writer.writerow(['Adresse', 'Kabellaengen', 'HA-KVz [m]'])
        # Get a list of all XLSX files in the specified directory
        filenames = glob.glob(path + "\\*.xlsx")
        # Iterate through each XLSX file
        for file in filenames:
            # Load the XLSX file using openpyxl
            wb = load_workbook(file, data_only=True)
            # Select the first worksheet in the workbook
            sh = wb.worksheets[0]
            # Get the values of specific cells
            adresse = sh['Z31'].value
            laenge = sh['BT45'].value
            # Write the values to the CSV file
            OTDR_writer.writerow([adresse, laenge])
            # Increment the address counter
            num_adress += 1
    # Convert all CSV files in the directory to XLSX format
    filenames = glob.glob(path + "\\*.csv")
    for file in filenames:
        read_file = pd.read_csv(file, encoding='latin-1')
        read_file.to_excel('OTDR_Kabellaenge' + '.xlsx', index=None, header=True)
    # Remove the temporary CSV file
    os.remove('OTDR.csv')
    return num_adress

def attenuation(path):
    num_invalid = 0
    # cells where the attenuation values are located
    cells_1310 = ['P87', 'P92', 'P97', 'P102', 'P107', 'P112', 'P117', 'BW87', 'BW92', 'BW97', 'BW102', 'BW107',
                  'BW112', 'BW117']
    cells_1550 = ['AN87', 'AN92', 'AN97', 'AN102', 'AN107', 'AN112', 'AN117', 'CU87', 'CU92', 'CU97', 'CU102',
                  'CU107', 'CU112', 'CU117']
    cells_1625 = ['AZ87', 'AZ92', 'AZ97', 'AZ102', 'AZ107', 'AZ112', 'AZ117', 'DG87', 'DG92', 'DG97', 'DG102',
                  'DG107', 'DGX12', 'DG117']
    # writing the adresses with too high attenuation values to a csv file
    with open('OTDR_Daempfung.csv', mode='w') as OTDR_file:
        writer = csv.writer(OTDR_file, lineterminator='\r')
        filenames = glob.glob(path + "\\*.xlsx")
        # checking every OTDR xlsx file
        for file in filenames:
            wb = load_workbook(file, data_only=True)
            sh = wb.worksheets[0]
            # amount of splices
            ns = sh['CY45'].value
            # cable length
            l = sh['BT45'].value
            # amount if grade B plugs
            nb = sh['CY50'].value
            # amount of grade C plugs
            nc = sh['CY55'].value
            invalid_value_found = False
            # Check each cell for invalid values and write the corresponding addresses to the CSV file
            for cell in cells_1310:
                try:
                    value = sh[cell].value
                    Daempfung_soll = 0.00036 * l + 0.2 * ns + 0.45 * nb + 0.7 * nc + 0.75
                    if value > Daempfung_soll:
                        invalid_value_found = True
                        writer.writerow([str(sh['Z31'].value)])
                        num_invalid += 1
                        break
                except TypeError:
                    continue
            if not invalid_value_found:
                for cell in cells_1550:
                    try:
                        value = sh[cell].value
                        Daempfung_soll = 0.00021 * l + 0.2 * ns + 0.45 * nb + 0.7 * nc + 0.75
                        if value > Daempfung_soll:
                            invalid_value_found = True
                            writer.writerow([str(sh['Z31'].value)])
                            num_invalid += 1
                            break
                    except TypeError:
                        continue
            if not invalid_value_found:
                for cell in cells_1625:
                    try:
                        value = sh[cell].value
                        Daempfung_soll = 0.00025 * l + 0.2 * ns + 0.45 * nb + 0.7 * nc + 0.75
                        if value > Daempfung_soll:
                            writer.writerow([str(sh['Z31'].value)])
                            num_invalid += 1
                            break
                    except TypeError:
                        continue
    return num_invalid

def print_result(adressen, invalid):
    # Return a formatted string with the number of addresses and the number of invalid values found
    return f'Fertig! Es wurden die Kabellängen von {adressen} ausgelesen. Bei {invalid} Adressen war mindestens ein Dämpfungswert zu hoch.'

if __name__ == '__main__':
    # Call the main function when the script is executed
    main()
